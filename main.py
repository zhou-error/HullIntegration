"""
船舶静水力数值积分计算主程序
基于参考文献：胡铁牛. 船舶静力学[M]. 上海交通大学出版社， 2022： 10-21.

设计6种计算方案：
  方案一：梯形法（基准方案，无端点修正，无半站）
  方案二：梯形法 + 端点线性修正
  方案三：辛普生第一法 + 端点线性修正
  方案四：辛普生第一法 + 端点抛物线修正 + 增加半站法
  方案五：辛普生第二法 + 端点线性修正
  方案六：乞贝雪夫法（9节点）

对每种方案计算：
  1. 各水线面面积 Aw
  2. 漂心纵坐标 xF
  3. 水线面横向惯性矩 IT
  4. 水线面纵向惯性矩 IL
  5. 排水体积曲线 ∇
  6. 浮心坐标 (LCB, KB)
  7. 船型系数 (CB, CM, CP, CWP)
"""

import os
import sys
import numpy as np
from datetime import datetime

from ship_data import load_offset_table, fill_none_with_zero, get_valid_station_range
from integration import (
    trapezoidal, simpson_13, simpson_38, chebyshev_discrete,
    endpoint_correction_linear, endpoint_correction_parabolic,
    simpson_13_weights, simpson_38_weights, trapezoidal_with_weights
)
from hydrostatics import HydrostaticCalculator


# ============================================================
#  配置
# ============================================================
OFFSET_TABLE_PATH = r'F:\claude\HullIntergration\OffsetTable\OffsetTable.xlsx'
OUTPUT_DIR = r'F:\claude\HullIntergration\Reports'

# ============================================================
#  方案定义
# ============================================================
SCHEMES = {
    '方案一': {
        'method': 'trapezoidal',
        'use_endpoint_correction': False,
        'use_half_stations': False,
        'endpoint_type': None,
        'description': '梯形法（基准方案，无端点修正，无增加半站）',
        'vertical_method': 'trapezoidal',
    },
    '方案二': {
        'method': 'trapezoidal',
        'use_endpoint_correction': True,
        'use_half_stations': False,
        'endpoint_type': 'linear',
        'description': '梯形法 + 端点线性修正',
        'vertical_method': 'trapezoidal',
    },
    '方案三': {
        'method': 'simpson13',
        'use_endpoint_correction': True,
        'use_half_stations': False,
        'endpoint_type': 'linear',
        'description': '辛普生第一法 + 端点线性修正',
        'vertical_method': 'trapezoidal',
    },
    '方案四': {
        'method': 'simpson13',
        'use_endpoint_correction': True,
        'use_half_stations': True,
        'endpoint_type': 'parabolic',
        'description': '辛普生第一法 + 端点抛物线修正 + 增加半站法',
        'vertical_method': 'trapezoidal',
    },
    '方案五': {
        'method': 'simpson38',
        'use_endpoint_correction': True,
        'use_half_stations': False,
        'endpoint_type': 'linear',
        'description': '辛普生第二法 + 端点线性修正',
        'vertical_method': 'trapezoidal',
    },
    '方案六': {
        'method': 'chebyshev',
        'use_endpoint_correction': False,
        'use_half_stations': False,
        'endpoint_type': None,
        'description': '乞贝雪夫法（9节点）',
        'vertical_method': 'trapezoidal',
    },
}


def run_all_schemes(data):
    """运行所有6种方案，返回结果字典"""
    calc = HydrostaticCalculator(data)
    all_results = {}

    for scheme_name, config in SCHEMES.items():
        print(f"  正在计算 {scheme_name}: {config['description']}...")
        results = run_scheme(calc, data, scheme_name, config)
        all_results[scheme_name] = results

    return all_results


def run_scheme(calc, data, scheme_name, config):
    """运行单个方案"""
    method = config['method']
    use_ec = config['use_endpoint_correction']
    use_hs = config['use_half_stations']
    ep_type = config['endpoint_type'] or 'linear'
    v_method = config['vertical_method']

    results = {
        'scheme_name': scheme_name,
        'description': config['description'],
        'waterline_results': [],
        'displacement_results': [],
        'sectional_areas': [],
        'process_tables': [],
        'method': method,
        'use_endpoint_correction': use_ec,
        'use_half_stations': use_hs,
        'endpoint_type': ep_type,
    }

    # --- 1. 各水线面计算 ---
    for wl_idx in range(1, data.num_waterlines):
        wl_label = data.waterline_labels[wl_idx]
        wl_height = data.waterline_heights[wl_idx]

        wl_result, process = calc.calc_waterplane_properties(
            wl_idx, method, use_ec, use_hs, ep_type
        )

        # 生成计算过程表
        table_data = calc.calc_waterplane_table_data(
            wl_idx, method, use_ec, use_hs
        )

        results['waterline_results'].append({
            'wl_label': wl_label,
            'wl_height': wl_height,
            'result': wl_result,
            'process': process,
            'table_data': table_data,
        })

    # --- 2. 排水体积及浮心 ---
    for wl_idx in range(1, data.num_waterlines):
        draft = data.waterline_heights[wl_idx]
        disp_result = calc.calc_displacement_properties(
            draft, method, use_ec, use_hs, v_method
        )
        results['displacement_results'].append({
            'wl_label': data.waterline_labels[wl_idx],
            'draft': draft,
            'result': disp_result,
        })

    return results


# ============================================================
#  报告生成
# ============================================================

def generate_reports(all_results, data):
    """为每种方案生成详细报告"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    dL = data.station_spacing or 1.0
    dL_label = "ΔL" if data.station_spacing is None else f"{dL:.0f}mm"

    for scheme_name, results in all_results.items():
        report_lines = []
        config = SCHEMES[scheme_name]

        # === 封面 ===
        report_lines.append("=" * 80)
        report_lines.append(f"      船舶静水力数值积分计算报告")
        report_lines.append(f"      {scheme_name}：{config['description']}")
        report_lines.append(f"      计算日期：{datetime.now().strftime('%Y年%m月%d日')}")
        report_lines.append("=" * 80)
        report_lines.append("")

        # === 1. 计算参数 ===
        report_lines.append("-" * 80)
        report_lines.append("1. 基本计算参数")
        report_lines.append("-" * 80)
        report_lines.append(f"  船舶站数：{data.num_stations}（站号0～{data.num_stations}）")
        report_lines.append(f"  站距：{dL_label}")
        report_lines.append(f"  水线数：{data.num_waterlines - 1}条")
        report_lines.append(f"  水线高度(mm)：{[int(h) for h in data.waterline_heights[1:]]}")
        report_lines.append("")

        # === 2. 水线面计算 ===
        report_lines.append("-" * 80)
        report_lines.append("2. 水线面要素计算")
        report_lines.append("-" * 80)

        for wr in results['waterline_results']:
            wl_label = wr['wl_label']
            wl_h = wr['wl_height']
            r = wr['result']
            tb = wr['table_data']

            report_lines.append(f"\n{'='*60}")
            report_lines.append(f"  水线：{wl_label} (高度 = {wl_h:.0f} mm)")
            report_lines.append(f"{'='*60}")

            # 2.1 计算表格
            report_lines.append(f"\n  2.1 半宽值及积分计算表")
            report_lines.append(f"  {'-'*56}")

            if 'weights' in tb:
                # 辛普生法表格
                report_lines.append(f"  {'站号':>6s}  {'半宽y(mm)':>10s}  {'辛普生乘数':>10s}  {'乘数×y':>12s}")
                report_lines.append(f"  {'-'*50}")
                for i in range(len(tb['station'])):
                    st = tb['station'][i]
                    y = tb['breadth'][i]
                    w = tb['weights'][i] if i < len(tb['weights']) else 0
                    wy = tb['weighted'][i] if i < len(tb['weighted']) else 0
                    report_lines.append(f"  {st:6.2f}  {y:10.1f}  {w:10.1f}  {wy:12.2f}")
                report_lines.append(f"  {'-'*50}")
                report_lines.append(f"  Σ(乘数×y) = {np.sum(tb['weighted']):.2f}")
                h_val = tb.get('h', 1.0)
                if method_label := tb.get('method_label', ''):
                    report_lines.append(f"  面积 = h/3 × Σ = {h_val:.4f} / 3 × {np.sum(tb['weighted']):.2f}")
                report_lines.append(f"  = {tb['total']:.4f}")

            elif 'segments' in tb:
                # 梯形法表格
                report_lines.append(f"  {'站号':>6s}  {'半宽y(mm)':>10s}  {'梯形面积分段':>14s}")
                report_lines.append(f"  {'-'*42}")
                for i in range(len(tb['station'])):
                    st = tb['station'][i]
                    y = tb['breadth'][i]
                    seg = tb['segments'][i] if i < len(tb['segments']) else 0
                    seg_str = f"{seg:14.4f}" if i < len(tb['segments']) else ""
                    report_lines.append(f"  {st:6.2f}  {y:10.1f}  {seg_str}")
                report_lines.append(f"  {'-'*42}")
                report_lines.append(f"  Σ(分段面积) = {tb['total']:.4f}")

            report_lines.append("")

            # 2.2 水线面要素汇总
            report_lines.append(f"  2.2 水线面静水力要素")
            report_lines.append(f"  {'-'*42}")
            report_lines.append(f"  半水线面积 ∫ydx    = {r['waterplane_area_half']:>12.4f} (×{dL_label})")
            report_lines.append(f"  水线面积 Aw=2∫ydx = {r['waterplane_area']:>12.4f} (×{dL_label})")
            report_lines.append(f"  对船中面积矩      = {r['moment_x_half']:>12.4f} (×{dL_label}²)")
            report_lines.append(f"  漂心距船中 xF     = {r['xf']:>12.4f} (站号)")
            report_lines.append(f"  半横向惯性矩∫y³dx= {r['IT_half']:>12.2f} (mm⁴)")
            report_lines.append(f"  横向惯性矩 IT     = {r['IT']:>12.2f} (mm⁴)")
            report_lines.append(f"  半纵向惯性矩（船中） = {r['IL_half_midship']:>12.2f} (×{dL_label}³·mm)")
            report_lines.append(f"  纵向惯性矩（船中） IL_mid = {2.0*r['IL_half_midship']:>12.2f} (×{dL_label}³·mm)")
            report_lines.append(f"  水线面系数 CWP    = {r['Cwp']:>12.6f}")
            report_lines.append("")

            # 端点修正信息
            ep_info = wr['process'].get('endpoint_info', {})
            if ep_info.get('corrected'):
                xr = ep_info.get('x_range', (0, 0))
                report_lines.append(f"  [端点修正] 水线实际范围: 站 {xr[0]:.3f} ~ {xr[1]:.3f}")
                report_lines.append(f"  [端点修正] 端点数: {ep_info.get('n_points', 0)}")
            report_lines.append("")

        # === 3. 排水体积计算 ===
        report_lines.append("-" * 80)
        report_lines.append("3. 排水体积及浮心计算")
        report_lines.append("-" * 80)

        for dr in results['displacement_results']:
            d = dr['draft']
            r = dr['result']
            report_lines.append(f"\n  --- 吃水 T = {d:.0f} mm ---")
            report_lines.append(f"  排水半体积 ∫Asdx   = {r['displacement_half']:>12.4f} (×{dL_label}³)")
            report_lines.append(f"  排水体积 ∇         = {r['displacement_volume']:>12.4f} (×{dL_label}³)")
            report_lines.append(f"  浮心距船中 LCB     = {r['LCB']:>12.4f} (站号)")
            report_lines.append(f"  浮心垂向 KB        = {r['KB']:>12.2f} (mm)")
            report_lines.append(f"  中横剖面面积 Am    = {r['Am']:>12.2f} (×{dL_label}²)")
            report_lines.append(f"  方形系数 CB        = {r['CB']:>12.6f}")
            report_lines.append(f"  中横剖面系数 CM    = {r['CM']:>12.6f}")
            report_lines.append(f"  水线面系数 CWP     = {r['CWP']:>12.6f}")
            report_lines.append(f"  菱形系数 CP        = {r['CP']:>12.6f}")
            report_lines.append("")

        # === 4. 总结 ===
        report_lines.append("-" * 80)
        report_lines.append("4. 方案特点总结")
        report_lines.append("-" * 80)
        report_lines.append(f"  积分方法：{results['method']}")
        report_lines.append(f"  端点修正：{'是 (' + results['endpoint_type'] + ')' if results['use_endpoint_correction'] else '否'}")
        report_lines.append(f"  增加半站：{'是' if results['use_half_stations'] else '否'}")
        report_lines.append("")

        # === 5. 参考文献 ===
        report_lines.append("-" * 80)
        report_lines.append("5. 参考文献")
        report_lines.append("-" * 80)
        report_lines.append("  [1] 胡铁牛. 船舶静力学[M]. 上海：上海交通大学出版社， 2022： 10-21.")
        report_lines.append("")

        # 写入文件
        safe_name = scheme_name.replace(' ', '_')
        report_path = os.path.join(OUTPUT_DIR, f'{safe_name}_报告.txt')
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(report_lines))
        print(f"    报告已保存: {report_path}")

    # === 汇总比较报告 ===
    generate_comparison_report(all_results, data)


def generate_comparison_report(all_results, data):
    """生成各方案汇总比较报告"""
    lines = []
    dL_label = "ΔL" if data.station_spacing is None else f"{data.station_spacing:.0f}mm"

    lines.append("=" * 80)
    lines.append("      各方案计算结果汇总比较")
    lines.append("=" * 80)
    lines.append("")

    # 取设计水线（WL5, 2000mm）作为比较基准
    design_wl_idx = 4  # WL5 = 2000mm (0-indexed among waterlines: 1→WL1, ..., 5→WL5)

    lines.append("-" * 80)
    lines.append(f"设计水线 (WL5, 2000mm) 水线面要素比较")
    lines.append("-" * 80)
    lines.append(f"{'方案':<12s} {'Aw(×dL)':<14s} {'xF(站号)':<12s} {'IT(mm⁴)':<16s} {'IL(×dL³·mm)':<18s} {'CWP':<10s}")
    lines.append("-" * 80)

    for scheme_name, results in all_results.items():
        if design_wl_idx < len(results['waterline_results']):
            wr = results['waterline_results'][design_wl_idx]
            r = wr['result']
            lines.append(
                f"{scheme_name:<12s} "
                f"{r['waterplane_area']:>12.4f}  "
                f"{r['xf']:>10.4f}  "
                f"{r['IT']:>14.2f}  "
                f"{2.0*r['IL_half']:>16.2f}  "
                f"{r['Cwp']:>8.6f}"
            )

    lines.append("")
    lines.append("-" * 80)
    lines.append(f"设计吃水 (2000mm) 排水体积及浮心比较")
    lines.append("-" * 80)
    lines.append(f"{'方案':<12s} {'∇(×dL³)':<14s} {'LCB(站号)':<12s} {'KB(mm)':<12s} {'CB':<10s} {'CM':<10s}")
    lines.append("-" * 80)

    for scheme_name, results in all_results.items():
        if design_wl_idx < len(results['displacement_results']):
            dr = results['displacement_results'][design_wl_idx]
            r = dr['result']
            lines.append(
                f"{scheme_name:<12s} "
                f"{r['displacement_volume']:>12.4f}  "
                f"{r['LCB']:>10.4f}  "
                f"{r['KB']:>10.2f}  "
                f"{r['CB']:>8.6f}  "
                f"{r['CM']:>8.6f}"
            )

    lines.append("")
    lines.append("-" * 80)
    lines.append("推荐结论")
    lines.append("-" * 80)
    lines.append("")
    lines.append("  综合考虑计算精度、稳定性和实用性，推荐使用：")
    lines.append("")
    lines.append("  ★ 方案四：辛普生第一法 + 端点抛物线修正 + 增加半站法 ★")
    lines.append("")
    lines.append("  推荐理由：")
    lines.append("  1. 辛普生第一法（1/3法则）基于二次抛物线插值，对船体型线这类光滑曲线")
    lines.append("     具有天然的计算优势，精度高于梯形法。")
    lines.append("  2. 端点抛物线修正能够更准确地确定水线首尾端点的实际位置，避免因站号")
    lines.append("     划分带来的截断误差。")
    lines.append("  3. 增加半站法在曲率变化较大的首尾部提供了更密集的数据点，显著提高了")
    lines.append("     积分精度。")
    lines.append("  4. 该方案在船舶静力学计算中应用最为广泛，方法成熟可靠，计算结果稳定。")
    lines.append("")
    lines.append("  注意事项：")
    lines.append("  1. 以上计算结果中的 x dL 表示需要乘以实际站距（mm）以获得绝对数值。")
    lines.append("  2. IT 和 IL 分别是水线面的横向和纵向惯性矩，IL 用于纵稳性计算。")
    lines.append("  3. 船型系数（CB, CM, CP, CWP）是无量纲量，不受站距影响。")
    lines.append("")

    # 参考文献
    lines.append("-" * 80)
    lines.append("参考文献")
    lines.append("-" * 80)
    lines.append("  [1] 胡铁牛. 船舶静力学[M]. 上海：上海交通大学出版社， 2022： 10-21.")
    lines.append("")

    report_path = os.path.join(OUTPUT_DIR, '各方案汇总比较报告.txt')
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f"    汇总报告已保存: {report_path}")


def generate_excel_report(all_results, data):
    """生成Excel格式的详细计算报告"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("    openpyxl不可用，跳过Excel报告生成")
        return

    wb = openpyxl.Workbook()

    # === Sheet 1: 汇总比较 ===
    ws_summary = wb.active
    ws_summary.title = "汇总比较"

    # 标题
    ws_summary.merge_cells('A1:J1')
    ws_summary['A1'] = '船舶静水力数值积分计算 — 各方案汇总比较'
    ws_summary['A1'].font = Font(bold=True, size=14)

    # 方案列表
    scheme_names = list(SCHEMES.keys())

    # 水线面要素比较（设计水线 WL5=2000mm）
    design_wl_idx = 4
    ws_summary['A3'] = '设计水线 (WL5, 2000mm) 水线面要素比较'
    ws_summary['A3'].font = Font(bold=True)

    headers = ['方案', '积分方法', '端点修正', '半站', 'Aw(×dL)', 'xF(站号)', 'IT(mm⁴)', 'IL(×dL³·mm)', 'CWP']
    for j, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=4, column=j, value=h)
        cell.font = Font(bold=True)

    for i, sn in enumerate(scheme_names):
        cfg = SCHEMES[sn]
        results = all_results[sn]
        if design_wl_idx < len(results['waterline_results']):
            wr = results['waterline_results'][design_wl_idx]
            r = wr['result']
            row = 5 + i
            ws_summary.cell(row=row, column=1, value=sn)
            ws_summary.cell(row=row, column=2, value=cfg['method'])
            ws_summary.cell(row=row, column=3, value=cfg['endpoint_type'] or '无')
            ws_summary.cell(row=row, column=4, value='是' if cfg['use_half_stations'] else '否')
            ws_summary.cell(row=row, column=5, value=round(r['waterplane_area'], 4))
            ws_summary.cell(row=row, column=6, value=round(r['xf'], 4))
            ws_summary.cell(row=row, column=7, value=round(r['IT'], 2))
            ws_summary.cell(row=row, column=8, value=round(2.0 * r['IL_half'], 2))
            ws_summary.cell(row=row, column=9, value=round(r['Cwp'], 6))

    # 排水体积比较
    start_row = 5 + len(scheme_names) + 2
    ws_summary.cell(row=start_row, column=1, value='设计吃水 (2000mm) 排水体积及浮心比较')
    ws_summary.cell(row=start_row, column=1).font = Font(bold=True)

    headers2 = ['方案', '∇(×dL³)', 'LCB(站号)', 'KB(mm)', 'CB', 'CM', 'CWP', 'CP']
    for j, h in enumerate(headers2, 1):
        ws_summary.cell(row=start_row + 1, column=j, value=h).font = Font(bold=True)

    for i, sn in enumerate(scheme_names):
        results = all_results[sn]
        if design_wl_idx < len(results['displacement_results']):
            dr = results['displacement_results'][design_wl_idx]
            r = dr['result']
            row = start_row + 2 + i
            ws_summary.cell(row=row, column=1, value=sn)
            ws_summary.cell(row=row, column=2, value=round(r['displacement_volume'], 4))
            ws_summary.cell(row=row, column=3, value=round(r['LCB'], 4))
            ws_summary.cell(row=row, column=4, value=round(r['KB'], 2))
            ws_summary.cell(row=row, column=5, value=round(r['CB'], 6))
            ws_summary.cell(row=row, column=6, value=round(r['CM'], 6))
            ws_summary.cell(row=row, column=7, value=round(r['CWP'], 6))
            ws_summary.cell(row=row, column=8, value=round(r['CP'], 6))

    # 推荐结论
    rec_row = start_row + 2 + len(scheme_names) + 2
    ws_summary.merge_cells(f'A{rec_row}:J{rec_row}')
    ws_summary.cell(row=rec_row, column=1, value='★ 推荐方案：方案四（辛普生第一法 + 端点抛物线修正 + 增加半站法）★').font = Font(bold=True, color='0000FF')

    # === Sheet 2: 方案四详细计算表 ===
    _write_detailed_sheet(wb, all_results, data, '方案四')

    # 调整列宽
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 25)

    xlsx_path = os.path.join(OUTPUT_DIR, '静水力计算汇总.xlsx')
    wb.save(xlsx_path)
    print(f"    Excel报告已保存: {xlsx_path}")


def _write_detailed_sheet(wb, all_results, data, scheme_name):
    """为指定方案写入详细计算表"""
    from openpyxl.styles import Font

    ws = wb.create_sheet(title=scheme_name)
    results = all_results[scheme_name]

    ws.merge_cells('A1:H1')
    ws['A1'] = f'{scheme_name} 详细计算过程'
    ws['A1'].font = Font(bold=True, size=12)
    ws['A2'] = f'积分方法: {results["method"]} | 端点修正: {results["endpoint_type"]} | 半站: {"是" if results["use_half_stations"] else "否"}'

    row = 4
    for wr in results['waterline_results'][:5]:  # 前5条水线
        wl_label = wr['wl_label']
        wl_h = wr['wl_height']
        r = wr['result']
        tb = wr['table_data']

        ws.cell(row=row, column=1, value=f'水线: {wl_label} (h={wl_h:.0f}mm)').font = Font(bold=True)
        row += 1

        # 表头
        headers = ['站号', '半宽y(mm)', '乘数', '乘数×y']
        for j, h in enumerate(headers, 1):
            ws.cell(row=row, column=j, value=h).font = Font(bold=True)
        row += 1

        if 'weights' in tb:
            for i in range(len(tb['station'])):
                ws.cell(row=row, column=1, value=tb['station'][i])
                ws.cell(row=row, column=2, value=tb['breadth'][i])
                ws.cell(row=row, column=3, value=tb['weights'][i] if i < len(tb['weights']) else 0)
                ws.cell(row=row, column=4, value=tb['weighted'][i] if i < len(tb['weighted']) else 0)
                row += 1
            ws.cell(row=row, column=1, value='Σ').font = Font(bold=True)
            ws.cell(row=row, column=4, value=np.sum(tb['weighted'])).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=1, value=f'面积 = h/3×Σ = {tb["total"]:.4f}')
            row += 2
        else:
            for i in range(len(tb['station'])):
                ws.cell(row=row, column=1, value=tb['station'][i])
                ws.cell(row=row, column=2, value=tb['breadth'][i])
                row += 1

        # 结果
        ws.cell(row=row, column=1, value=f'Aw = {r["waterplane_area"]:.4f}  xF = {r["xf"]:.4f}  IT = {r["IT"]:.2f}')
        row += 2


# ============================================================
#  主程序
# ============================================================

def main():
    print("=" * 60)
    print("  船舶静水力数值积分计算程序")
    print("  参考：胡铁牛《船舶静力学》2022，pp.10-21")
    print("=" * 60)
    print()

    # 1. 加载数据
    print("1. 加载型值表...")
    data = load_offset_table(OFFSET_TABLE_PATH)
    print(f"   读取到 {len(data.stations)} 个站位（含半站）")
    print(f"   水线数: {data.num_waterlines - 1}")
    print(f"   水线高度: {[int(h) for h in data.waterline_heights[1:]]} mm")
    print()

    # 2. 运行所有方案
    print("2. 运行6种计算方案...")
    all_results = run_all_schemes(data)
    print()

    # 3. 生成报告
    print("3. 生成计算报告...")
    generate_reports(all_results, data)
    print()

    # 4. 生成Excel报告
    print("4. 生成Excel汇总报告...")
    generate_excel_report(all_results, data)
    print()

    print("=" * 60)
    print("  计算完成！报告保存在 Reports/ 文件夹下。")
    print("=" * 60)


if __name__ == '__main__':
    main()
