"""
船舶型值表数据加载模块
读取OffsetTable.xlsx，解析站号、水线高度、半宽值等数据。
"""

import openpyxl
import numpy as np
from dataclasses import dataclass, field
from typing import List, Dict, Tuple, Optional


@dataclass
class ShipOffsetData:
    """船舶型值表数据结构"""
    # 基本参数
    num_stations: int = 20  # 等分站数（站距数）
    stations: List[float] = field(default_factory=list)  # 站号列表 [-0.5, 0, 1, ..., 20]
    station_positions: List[int] = field(default_factory=list)  # 整数站号 [0,1,...,20]

    # 水线参数
    waterline_labels: List[str] = field(default_factory=list)  # ['基线','WL1','WL2',...]
    waterline_heights: List[float] = field(default_factory=list)  # [0, 400, 800, ...]
    num_waterlines: int = 0

    # 半宽值数据: half_breadths[station][wl_index] = 半宽值(mm) 或 None(无线型)
    half_breadths: Dict[float, List[Optional[float]]] = field(default_factory=dict)

    # 甲板相关数据
    deck_side: Dict[float, Optional[float]] = field(default_factory=dict)
    deck_center: Dict[float, Optional[float]] = field(default_factory=dict)
    deck_side_height: Dict[float, Optional[float]] = field(default_factory=dict)
    bulwark_top: Dict[float, Optional[float]] = field(default_factory=dict)
    fc_deck_side: Dict[float, Optional[float]] = field(default_factory=dict)
    poop_deck_side: Dict[float, Optional[float]] = field(default_factory=dict)
    fc_deck_center: Dict[float, Optional[float]] = field(default_factory=dict)
    poop_deck_center: Dict[float, Optional[float]] = field(default_factory=dict)

    # 站距（需用户提供或根据船长计算）
    station_spacing: float = None  # dL, 单位mm

    def set_station_spacing(self, Lpp_mm: float):
        """根据垂线间长设置站距"""
        self.station_spacing = Lpp_mm / self.num_stations

    def get_half_breadth(self, station: float, wl_idx: int) -> Optional[float]:
        """获取指定站号和水线索引的半宽值"""
        if station in self.half_breadths:
            vals = self.half_breadths[station]
            if wl_idx < len(vals):
                return vals[wl_idx]
        return None


def load_offset_table(filepath: str) -> ShipOffsetData:
    """从Excel文件加载船舶型值表"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Sheet1']

    data = ShipOffsetData()

    # --- 解析头部信息 ---
    # Row 1: 站号, 型值, ..., 高度, ..., 站号
    # Row 2: None, WL1, WL2, ..., WL8, 甲板边线, 舷墙顶线, ..., 甲板中线, None
    # Row 3: None, 400, 800, ..., 3200 (水线高度，mm), None, None, 400, 800, ... (甲板高度)

    # 读取水线标签和高度
    waterline_heights_raw = []
    for col in range(2, 11):  # B-J columns (1-indexed: columns 2-10)
        label = ws.cell(row=2, column=col).value
        height = ws.cell(row=3, column=col).value
        if label and height is not None:
            waterline_heights_raw.append((str(label), float(height)))

    data.waterline_labels = [f"基线"] + [wl[0] for wl in waterline_heights_raw]
    data.waterline_heights = [0.0] + [wl[1] for wl in waterline_heights_raw]
    data.num_waterlines = len(data.waterline_labels)

    # --- 解析站号和各站半宽值 ---
    # 站号在A列（column 1），数据从row 4开始
    data.stations = []
    data.station_positions = []
    data.half_breadths = {}

    for row_idx in range(4, 25):  # Rows 4-24: 站 -0.5 到 20
        station_val = ws.cell(row=row_idx, column=1).value  # A列：站号
        if station_val is None:
            continue

        # 处理站号
        try:
            station = float(station_val)
        except (ValueError, TypeError):
            # 可能是"艉封板"等文字
            continue

        data.stations.append(station)
        if station == int(station) and station >= 0:
            data.station_positions.append(int(station))

        # 读取半宽值（B列：型值/基线，C-J列：WL1-WL8）
        half_breadths = []
        for col in range(2, 11):  # B-J columns
            val = ws.cell(row=row_idx, column=col).value
            if val is None or val == '\\':
                half_breadths.append(None)
            else:
                try:
                    half_breadths.append(float(val))
                except (ValueError, TypeError):
                    # 处理 "397/1063/1768" 这种多值格式，取第一个
                    s = str(val).split('/')[0]
                    try:
                        half_breadths.append(float(s))
                    except ValueError:
                        half_breadths.append(None)

        data.half_breadths[station] = half_breadths

    # --- 解析甲板数据 ---
    # K列(col 11): 甲板边线, L列(col 12): 舷墙顶线, M(col 13): 首楼甲板边线
    # N(col 14): 尾楼甲板边线, O(col 15): 首楼甲板中线, P(col 16): 尾楼甲板中线
    # Q(col 17): 甲板边线高度, R(col 18): 甲板中线

    for row_idx in range(4, 25):
        station_val = ws.cell(row=row_idx, column=1).value
        if station_val is None:
            continue
        try:
            station = float(station_val)
        except (ValueError, TypeError):
            continue

        def _safe_float(v):
            if v is None or v == '\\':
                return None
            try:
                return float(v)
            except (ValueError, TypeError):
                s = str(v).split('/')[0]
                try:
                    return float(s)
                except ValueError:
                    return None

        data.deck_side[station] = _safe_float(ws.cell(row=row_idx, column=11).value)
        data.bulwark_top[station] = _safe_float(ws.cell(row=row_idx, column=12).value)
        data.fc_deck_side[station] = _safe_float(ws.cell(row=row_idx, column=13).value)
        data.poop_deck_side[station] = _safe_float(ws.cell(row=row_idx, column=14).value)
        data.fc_deck_center[station] = _safe_float(ws.cell(row=row_idx, column=15).value)
        data.poop_deck_center[station] = _safe_float(ws.cell(row=row_idx, column=16).value)
        data.deck_side_height[station] = _safe_float(ws.cell(row=row_idx, column=17).value)
        data.deck_center[station] = _safe_float(ws.cell(row=row_idx, column=18).value)

    # 排序站号
    data.stations.sort()
    data.station_positions.sort()

    return data


def get_valid_station_range(data: ShipOffsetData, wl_idx: int) -> Tuple[float, float, int, int]:
    """
    获取某水线有效范围：找到第一个和最后一个非None半宽值对应的站号及索引。
    用于端点修正。
    返回: (first_station, last_station, first_idx, last_idx)
    """
    stations = data.stations
    first_idx = None
    last_idx = None

    for i, s in enumerate(stations):
        hb = data.get_half_breadth(s, wl_idx)
        if hb is not None:
            if first_idx is None:
                first_idx = i
            last_idx = i

    if first_idx is None:
        return None, None, None, None

    return stations[first_idx], stations[last_idx], first_idx, last_idx


def interpolate_half_stations(data: ShipOffsetData, wl_idx: int) -> Tuple[np.ndarray, np.ndarray]:
    """
    在整数站之间插入半站，提高计算精度（增加半站法）。
    对于有None值的水线，在半站位置进行线性插值。
    返回: (expanded_stations, expanded_breadths)
    """
    stations = data.stations
    orig_breadths = []
    for s in stations:
        hb = data.get_half_breadth(s, wl_idx)
        orig_breadths.append(hb)

    # 构建扩展后的站号和半宽值
    expanded_stations = []
    expanded_breadths = []

    for i in range(len(stations) - 1):
        s1, s2 = stations[i], stations[i + 1]
        y1, y2 = orig_breadths[i], orig_breadths[i + 1]

        # 添加当前站
        expanded_stations.append(s1)
        expanded_breadths.append(y1)

        # 添加半站
        s_mid = (s1 + s2) / 2.0
        expanded_stations.append(s_mid)

        # 半站处线性插值
        if y1 is None and y2 is None:
            expanded_breadths.append(None)
        elif y1 is None:
            expanded_breadths.append(y2 * 0.5)  # 从0开始近似
        elif y2 is None:
            expanded_breadths.append(y1 * 0.5)
        else:
            expanded_breadths.append((y1 + y2) / 2.0)

    # 添加最后一个站
    expanded_stations.append(stations[-1])
    expanded_breadths.append(orig_breadths[-1])

    return np.array(expanded_stations), np.array(expanded_breadths, dtype=object)


def fill_none_with_zero(breadths: np.ndarray) -> np.ndarray:
    """将None值替换为0，用于数值积分"""
    return np.array([0.0 if y is None else float(y) for y in breadths])


def get_waterline_breadths(data: ShipOffsetData, wl_idx: int) -> Tuple[np.ndarray, np.ndarray]:
    """
    获取某水线的站号和半宽值数组。
    返回: (station_array, breadth_array) - breadth中None表示无线型
    """
    stations = np.array(data.stations)
    breadths = np.array([data.get_half_breadth(s, wl_idx) for s in data.stations], dtype=object)
    return stations, breadths
